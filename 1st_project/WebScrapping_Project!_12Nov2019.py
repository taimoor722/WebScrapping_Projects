import requests
from bs4 import BeautifulSoup
import pprint,time,datetime
no_change_day=12
no_change_month=11
no_change_year=2019
no_change_date=f'{no_change_day} {no_change_month}-{no_change_year}'
start=time.time()
today_date=datetime.datetime.now()
cutom_formated_todays_date=today_date.strftime('%d %b-%Y')
cutom_formated_todays_time=today_date.strftime('%I o\'clock %p')

#print(cutom_formated_todays_date)
run_one_time=0
My_task_date=(today_date.minute)
print('program will run on this date now-->',My_task_date)
print('but rigth now its---->',today_date.minute)
col_next=1
while True:
     today_date = datetime.datetime.now()
     time.sleep(2)
     print('continue waiting')
     print('program will run on this date now-->', My_task_date)
     print('but rigth now its---->', today_date.minute)

     while My_task_date==today_date.minute:#and run_one_time==0
        print('we are now in program')
        My_task_date=(today_date.minute)+3
        run_one_time += 1
        print(My_task_date)

        taken=requests.get('https://www.producthunt.com/')
        #print(taken.raise_for_status())
        #print(taken.status_code==requests.codes.ok)

        html=BeautifulSoup(taken.text,"html.parser")
        reached_anchor=html.find_all('a',"link_523b9")

        print('->->',len(reached_anchor),reached_anchor)
        print(reached_anchor)
        import webbrowser
        all_link_list=[]
        unique=[]
        for i in reached_anchor:
            i=i.get('href')

            print('this is i========>',i)
            link='https://www.producthunt.com'+i
            #if str(i).startswith('/r/p') or str(i).startswith('/r'):
            taken_unique=requests.get(link)
            html_unique=BeautifulSoup(taken_unique.text,'html.parser')
            #print('st',html_unique,'end')
            anchor_unique=html_unique.head.find_all('meta',{'property':'og:url'})
            print(len(anchor_unique),'*********************',[i for i in anchor_unique])
            print('###########',anchor_unique[0].get('content'))
            un=anchor_unique[0].get('content')
            if str(un).startswith('http'):
                   unique.append(un)

            print(link)
            taken1=requests.get(link)
            #webbrowser.open(link)
            #print(taken.raise_for_status())
            #print(taken.status_code == requests.codes.ok)
            print(taken1,i,'taken1')
            html1=BeautifulSoup(taken1.text,"html.parser")
            reached_anchor1=html1.find_all('a',["link_9bebc"])
            print(reached_anchor1)
            for k in reached_anchor1:

                print('k',k)
                link_final=k.get('title',1)
                print('lf',link_final)
                print(html1.title,len(reached_anchor1), reached_anchor1, 'pp',sep='\n')
                if 'www.'not in link_final:
                   link_final='www.'+link_final
                if link_final not in ['www.play.google.com','www.apps.apple.com']:
                   print('finfifinfinfin',link_final)
                   all_link_list.append(link_final)




        print('##uniq333',len(unique),unique)
        all_link_list.extend(unique)
        pprint.pprint(all_link_list)


        from openpyxl import load_workbook
        from openpyxl.styles import fonts,Font,colors,alignment,Alignment,fills,Fill,PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import fills,PatternFill
        excel=load_workbook('web_scrapping_project1_upwork.xlsx')
        sheet=excel.active
        print(sheet.title)

        #print(get_column_letter(col_next))
        sheet[f'{get_column_letter(idx=col_next)}1']=f'All links Date'
        sheet[f'{get_column_letter(idx=col_next)}1'].font=Font(bold=True,color=colors.RED,size=15)

        sheet[f'{get_column_letter(idx=col_next)}2']=f'{cutom_formated_todays_date}'
        sheet[f'{get_column_letter(idx=col_next)}2'].font=Font(bold=True,color=colors.RED,size=15)

        sheet[f'{get_column_letter(idx=col_next)}3']=f'{cutom_formated_todays_time}'
        sheet[f'{get_column_letter(idx=col_next)}3'].font=Font(bold=True,color=colors.RED,size=15)

        sheet[f'{get_column_letter(idx=col_next)}1'].alignment=Alignment(horizontal='center')
        sheet[f'{get_column_letter(idx=col_next)}2'].alignment = Alignment(horizontal='center')
        #for r in range(len(all_link_list)):
        sheet.column_dimensions[f'{get_column_letter(idx=col_next)}'].width=38
        sheet.row_dimensions[1].height=20
        sheet.row_dimensions[2].height=20
        e=4
        for row in all_link_list:
            e+=1
            sheet[f'{get_column_letter(idx=col_next)}{e}']=row
        #print(sheet[f'{get_column_letter(col_next)}'])
        col_next += 2
        col_to_color=col_next-1
        #sheet[f'{get_column_letter(col_to_color)}'].fills=PatternFill(bgColor=colors.GREEN,fill_type='solid')
        excel.save('web_scrapping_project1_upwork.xlsx')
        excel.close()

        import pandas as pd
        import numpy as np
        from pandas import ExcelFile,ExcelWriter
        df=pd.array(all_link_list)
        EW=ExcelWriter('web_scrapping_project1_upwork.xlsx')
        
        df.to_excel(EW,'sheet',index=False)
        EW.save()
        EW.close()
        stop = time.time()
        print('processing done in ' ,int(stop)-int(start), '..sec')
        #reached_anchor1=reached_anchor1[0]
        #reached_anchor1=reached_anchor1["link_9bebc"]
